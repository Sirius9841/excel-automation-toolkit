"""Tests for exporter.py — Excel and CSV export."""

import sys
from pathlib import Path
from io import BytesIO

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.exporter import (
    export_to_excel,
    export_to_csv,
    _safe_filename,
    ExportError,
)


# ── Fixtures ──────────────────────────────────────────────


@pytest.fixture
def df_simple() -> pd.DataFrame:
    return pd.DataFrame({
        "product": ["Widget A", "Widget B", "Gadget X"],
        "price": [10.50, 25.00, 99.99],
        "quantity": [5, 3, 1],
    })


@pytest.fixture
def df_with_special_chars() -> pd.DataFrame:
    return pd.DataFrame({
        "city": ["München", "São Paulo", "Côte d'Azur"],
        "value": [100, 200, 300],
    })


@pytest.fixture
def df_dates() -> pd.DataFrame:
    return pd.DataFrame({
        "date": pd.to_datetime(["2024-01-01", "2024-06-15", "2024-12-31"]),
        "amount": [1500.50, 2300.00, 4200.75],
    })


# ── Test: filename safety ─────────────────────────────────


class TestSafeFilename:

    def test_replaces_invalid_chars(self):
        result = _safe_filename('file<>:"/|?*name', ".csv")
        assert result == "file________name.csv"

    def test_replaces_spaces_with_underscores(self):
        result = _safe_filename("my cleaned data", ".xlsx")
        assert result == "my_cleaned_data.xlsx"

    def test_empty_base_uses_fallback(self):
        result = _safe_filename("   ", ".csv")
        assert result == "exported_data.csv"

    def test_keeps_valid_names(self):
        result = _safe_filename("cleaned_dataset_2024-07-28", ".xlsx")
        assert result == "cleaned_dataset_2024-07-28.xlsx"


# ── Test: Excel export ────────────────────────────────────


class TestExportToExcel:

    def test_returns_bytes_and_filename(self, df_simple):
        data, filename = export_to_excel(df_simple, "test_export")
        assert isinstance(data, bytes)
        assert filename.endswith(".xlsx")
        assert len(data) > 0

    def test_auto_filename_when_empty_base(self, df_simple):
        data, filename = export_to_excel(df_simple)
        assert filename.endswith(".xlsx")
        assert "cleaned_dataset" in filename

    def test_excel_is_readable(self, df_simple):
        data, _ = export_to_excel(df_simple, "test_roundtrip")
        roundtrip = pd.read_excel(BytesIO(data))
        assert list(roundtrip.columns) == ["Product", "Price", "Quantity"]
        pd.testing.assert_frame_equal(
            roundtrip.set_axis(df_simple.columns, axis=1),
            df_simple,
        )

    def test_preserves_special_characters(self, df_with_special_chars):
        data, _ = export_to_excel(df_with_special_chars, "special")
        roundtrip = pd.read_excel(BytesIO(data))
        assert roundtrip["City"].iloc[0] == "München"
        assert roundtrip["City"].iloc[1] == "São Paulo"

    def test_preserves_dates(self, df_dates):
        data, _ = export_to_excel(df_dates, "dates")
        roundtrip = pd.read_excel(BytesIO(data))
        pd.testing.assert_series_equal(
            roundtrip["Date"].dt.date.rename("date"),
            df_dates["date"].dt.date,
        )

    def test_mixed_missing_values_produce_readable_workbook(self):
        df = pd.DataFrame({
            "strings": ["Alpha", None, "Gamma"],
            "floats": [1.25, float("nan"), 3.75],
            "nullable": pd.Series([1, pd.NA, 3], dtype="Int64"),
            "timestamps": [
                pd.Timestamp("2024-01-01"),
                pd.NaT,
                pd.Timestamp("2024-03-15"),
            ],
            "mixed": pd.Series(["text", 2.5, None], dtype="object"),
        })

        data, filename = export_to_excel(df, "mixed_values")
        workbook = load_workbook(BytesIO(data))

        assert filename == "mixed_values.xlsx"
        assert "Cleaned Data" in workbook.sheetnames
        assert workbook["Cleaned Data"].max_row == 4
        assert workbook["Cleaned Data"].max_column == 5

    def test_empty_dataframe_raises_error(self):
        with pytest.raises(ExportError, match="empty"):
            export_to_excel(pd.DataFrame(), "empty")

    def test_dataframe_not_modified(self, df_simple):
        original = df_simple.copy()
        export_to_excel(df_simple, "immutable")
        pd.testing.assert_frame_equal(df_simple, original)


# ── Test: CSV export ──────────────────────────────────────


class TestExportToCSV:

    def test_returns_bytes_and_filename(self, df_simple):
        data, filename = export_to_csv(df_simple, "test_export")
        assert isinstance(data, bytes)
        assert filename.endswith(".csv")
        assert len(data) > 0

    def test_auto_filename_when_empty_base(self, df_simple):
        data, filename = export_to_csv(df_simple)
        assert filename.endswith(".csv")
        assert "cleaned_dataset" in filename

    def test_csv_is_readable(self, df_simple):
        data, _ = export_to_csv(df_simple, "test_roundtrip")
        roundtrip = pd.read_csv(BytesIO(data))
        pd.testing.assert_frame_equal(roundtrip, df_simple)

    def test_utf8_encoding(self, df_with_special_chars):
        data, _ = export_to_csv(df_with_special_chars, "utf8_test")
        content = data.decode("utf-8-sig")
        assert "München" in content
        assert "São Paulo" in content

    def test_empty_dataframe_raises_error(self):
        with pytest.raises(ExportError, match="empty"):
            export_to_csv(pd.DataFrame(), "empty")

    def test_dataframe_not_modified(self, df_simple):
        original = df_simple.copy()
        export_to_csv(df_simple, "immutable")
        pd.testing.assert_frame_equal(df_simple, original)
