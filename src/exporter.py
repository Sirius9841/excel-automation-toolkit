"""Create business-ready Excel and interoperable CSV exports in memory."""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from math import ceil
from typing import Any, Iterable, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.analyzer import detect_business_column_type, friendly_column_name
from src.data_quality import (
    CleaningAuditEntry,
    audit_records,
    business_action_text,
    classify_missing_values,
    missing_status_by_column,
    structural_blank_audit,
)
from src.logger_setup import setup_logger

logger = setup_logger(__name__)

EXCEL_SHEET_NAME = "Cleaned Data"
SUMMARY_SHEET_NAME = "Cleaning Summary"
REVIEW_SHEET_NAME = "Values to Review"
AUDIT_SHEET_NAME = "Cleaning Audit"
MAX_COL_WIDTH = 42

NAVY = "0F172A"
BLUE = "2563EB"
BLUE_GRAY = "E8EEF5"
LIGHT_BLUE = "F4F7FC"
MID_BORDER = "CBD5E1"
MUTED = "475569"
WHITE = "FFFFFF"
GREEN = "15803D"
AMBER = "B45309"


class ExportError(Exception):
    """Raised when export preparation fails."""


def _safe_filename(base: str, extension: str) -> str:
    """Build a filesystem-safe suggested filename."""
    safe = re.sub(r'[<>:"/\\|?*]', "_", base.strip())
    safe = re.sub(r"\s+", "_", safe)
    if not safe:
        safe = "exported_data"
    return f"{safe}{extension}"


def _is_missing(value: object) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _display_length(value: object) -> int:
    """Return a safe display length for mixed spreadsheet values."""
    if _is_missing(value):
        return 0
    if isinstance(value, (pd.Timestamp, datetime)):
        return 10
    return len(str(value))


def _is_identifier_column(column: object, series: pd.Series) -> bool:
    return detect_business_column_type(series.rename(column)) == "identifier"


def _clean_scalar(value: object) -> object:
    if _is_missing(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    return value


def _prepare_export_dataframe(df: pd.DataFrame, *, for_csv: bool) -> pd.DataFrame:
    """Normalize values consistently without modifying the source DataFrame."""
    prepared = df.copy(deep=True)
    for column in prepared.columns:
        series = prepared[column]
        if _is_identifier_column(column, series):
            prepared[column] = series.map(
                lambda value: None
                if _is_missing(value)
                else str(int(value))
                if isinstance(value, float) and value.is_integer()
                else str(value)
            )
            continue
        if pd.api.types.is_datetime64_any_dtype(series):
            if for_csv:
                prepared[column] = series.map(
                    lambda value: ""
                    if _is_missing(value)
                    else pd.Timestamp(value).strftime("%Y-%m-%d")
                )
            continue
        if (
            pd.api.types.is_numeric_dtype(series)
            and not pd.api.types.is_bool_dtype(series)
        ):
            prepared[column] = pd.Series(
                [
                    None
                    if _is_missing(value)
                    else int(value)
                    if float(value).is_integer()
                    else float(value)
                    for value in series
                ],
                index=series.index,
                dtype=object,
            )
    return prepared.astype(object).where(pd.notna(prepared), None)


def _header_style(cell: Any) -> None:
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color=MID_BORDER))


def _style_table_sheet(
    worksheet: Any,
    *,
    freeze: str = "A2",
    autofilter_table_name: str | None = None,
    header_row: int = 1,
) -> None:
    worksheet.freeze_panes = freeze
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[header_row].height = 24
    for cell in worksheet[header_row]:
        _header_style(cell)

    for row in worksheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.font = Font(name="Calibri", size=10, color=NAVY)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        worksheet.row_dimensions[row[0].row].height = 20

    for column_index, column_cells in enumerate(
        worksheet.iter_cols(1, worksheet.max_column),
        start=1,
    ):
        max_length = max(
            (_display_length(cell.value) for cell in column_cells),
            default=0,
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 12),
            MAX_COL_WIDTH,
        )

    if autofilter_table_name and worksheet.max_row >= 2 and worksheet.max_column:
        reference = (
            f"A{header_row}:"
            f"{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        )
        table = Table(displayName=autofilter_table_name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)


def _fit_wrapped_row_heights(worksheet: Any) -> None:
    """Increase row heights for wrapped narrative cells."""
    for row_index in range(2, worksheet.max_row + 1):
        line_count = 1
        for column_index in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            width = (
                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width
                or 12
            )
            estimated_chars = max(int(width * 1.45), 8)
            cell_lines = sum(
                max(1, ceil(len(part) / estimated_chars))
                for part in str(value).splitlines()
            )
            line_count = max(line_count, cell_lines)
        worksheet.row_dimensions[row_index].height = max(
            20,
            15 * line_count + 4,
        )


def _write_dataframe(
    worksheet: Any,
    df: pd.DataFrame,
    *,
    friendly_headers: bool = False,
) -> None:
    headers = [
        friendly_column_name(column) if friendly_headers else str(column)
        for column in df.columns
    ]
    worksheet.append(headers)
    for row in df.itertuples(index=False, name=None):
        worksheet.append([_clean_scalar(value) for value in row])


def _apply_cleaned_data_formats(worksheet: Any, df: pd.DataFrame) -> None:
    for column_index, column in enumerate(df.columns, start=1):
        series = df[column]
        business_type = detect_business_column_type(series.rename(column))
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.value is None:
                continue
            if business_type == "identifier":
                cell.number_format = "@"
                cell.alignment = Alignment(vertical="center", horizontal="left")
            elif pd.api.types.is_datetime64_any_dtype(series):
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(vertical="center", horizontal="center")
            elif (
                pd.api.types.is_integer_dtype(series)
                and not pd.api.types.is_bool_dtype(series)
            ):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(vertical="center", horizontal="right")
            elif (
                pd.api.types.is_numeric_dtype(series)
                and not pd.api.types.is_bool_dtype(series)
            ):
                cell.number_format = (
                    "#,##0"
                    if float(cell.value).is_integer()
                    else "#,##0.00"
                )
                cell.alignment = Alignment(vertical="center", horizontal="right")


def _summary_rows(
    df: pd.DataFrame,
    cleaning_report: Mapping[str, Any] | None,
    source_files: list[str],
) -> list[tuple[str, object, str]]:
    report = dict(cleaning_report or {})
    rows = [
        ("Report generated", datetime.now().strftime("%Y-%m-%d %H:%M"), ""),
        ("Source files", ", ".join(source_files) if source_files else "—", ""),
        ("Rows before cleaning", report.get("rows_before", len(df)), ""),
        ("Rows after cleaning", report.get("rows_after", len(df)), ""),
        (
            "Duplicate rows removed",
            report.get("duplicates_removed", 0),
            "Repeated records removed after user approval.",
        ),
        (
            "Incomplete rows removed",
            report.get("incomplete_rows_removed", 0),
            "Rows removed because source-provided values were blank.",
        ),
        (
            "Missing values reviewed",
            report.get("missing_values_reviewed", 0),
            "Source-provided fields reviewed through an approved decision.",
        ),
        (
            "Values changed",
            report.get("values_changed", report.get("values_filled", 0)),
            "Values recovered, filled, or removed through an approved action.",
        ),
        (
            "Approved to remain blank",
            report.get("approved_unchanged", 0),
            "Physical blanks retained after explicit user approval.",
        ),
        (
            "Decisions pending",
            report.get("decisions_pending", 0),
            "Missing-value decisions that still require an approved action.",
        ),
        (
            "Unavailable from source",
            report.get(
                "unavailable_from_source",
                report.get("structural_missing_after", 0),
            ),
            "The original source file did not contain the field.",
        ),
        (
            "Integrity failures",
            report.get(
                "integrity_failures",
                report.get("severe_integrity_issue_count", 0),
            ),
            "Validated relationship checks that require attention.",
        ),
    ]
    return rows


def _write_cleaning_summary(
    worksheet: Any,
    df: pd.DataFrame,
    cleaning_report: Mapping[str, Any] | None,
    source_files: list[str],
    source_schemas: Mapping[str, Iterable[str]] | None,
    cleaning_audit: Iterable[CleaningAuditEntry | Mapping[str, Any]] | None,
) -> None:
    worksheet.append(["Metric", "Value", "Explanation"])
    for row in _summary_rows(df, cleaning_report, source_files):
        worksheet.append(list(row))

    report = dict(cleaning_report or {})
    actions = [business_action_text(action) for action in report.get("missing_actions", [])]
    severe_issue_count = int(report.get("severe_integrity_issue_count", 0))
    unresolved = (
        [
            (
                f"{severe_issue_count:,} severe relationship "
                f"{'issue requires' if severe_issue_count == 1 else 'issues require'} "
                "business review."
            )
        ]
        if severe_issue_count
        else ["No unresolved integrity problems were found."]
    )
    sections = [
        (
            "Actions applied",
            actions or ["No cleaning actions were applied."],
            "See Cleaning Audit",
        ),
        (
            "Unavailable source fields",
            [
                (
                    f"{report.get('structural_missing_after', 0):,} cells were "
                    "kept blank because their source files did not include "
                    "those fields."
                )
            ],
            "No action required",
        ),
        (
            "Estimated values",
            [
                f"{report.get('estimated_values', 0):,} values were estimated."
            ],
            "Review calculation scope in Cleaning Audit",
        ),
        (
            "Deterministic recoveries",
            [
                (
                    f"{report.get('deterministic_recoveries', 0):,} values were "
                    "recovered from validated arithmetic relationships."
                )
            ],
            "Formula and inputs are recorded in Cleaning Audit",
        ),
        (
            "Unresolved problems",
            unresolved,
            (
                "Acknowledge before operational use"
                if severe_issue_count
                else "No action required"
            ),
        ),
    ]
    for section, details, next_step in sections:
        worksheet.append([])
        worksheet.append([section, "Details", "Next step"])
        for index, detail in enumerate(details, start=1):
            item_label = (
                f"Action {index}"
                if section == "Actions applied"
                else "Summary"
            )
            worksheet.append([item_label, detail, next_step])

    worksheet.append([])
    worksheet.append(["Integrity Checks", "Result", "Details"])
    integrity_passed = bool(report.get("integrity_passed", True))
    worksheet.append([
        "Relationship validation",
        "Passed" if integrity_passed else "Failed",
        (
            "All complete records satisfy the validated relationships."
            if integrity_passed
            else (
                f"{report.get('severe_integrity_issue_count', 0):,} severe "
                "relationship issue(s) require review."
            )
        ),
    ])

    status_by_column = missing_status_by_column(
        df,
        source_schemas,
        cleaning_audit or (),
    )
    status_rows = [
        (
            friendly_column_name(column),
            status.approved_blank,
            status.unavailable_from_source,
            status.decisions_pending,
        )
        for column, status in status_by_column.items()
        if (
            status.approved_blank
            or status.unavailable_from_source
            or status.decisions_pending
        )
    ]
    if status_rows:
        worksheet.append([])
        worksheet.append([
            "Blank status by field",
            "Approved blank",
            "Unavailable from source",
            "Decisions pending",
        ])
        for row in status_rows:
            worksheet.append(list(row))


def _review_export_frame(outlier_df: pd.DataFrame | None) -> pd.DataFrame:
    columns = [
        "Record ID",
        "Original Source Row",
        "Source File",
        "Product",
        "Field",
        "Value",
        "Upper Review Threshold",
    ]
    if outlier_df is None or outlier_df.empty:
        return pd.DataFrame(columns=columns)
    result = outlier_df.copy().rename(columns={
        "Description": "Product",
        "Column": "Field",
        "Upper Review Boundary": "Upper Review Threshold",
    })
    if "Review Type" in result.columns and "Product" in result.columns:
        integrity_mask = result["Review Type"].eq("Integrity check")
        result.loc[integrity_mask, "Product"] = (
            "Integrity check — Relationship check failed — "
            + result.loc[integrity_mask, "Product"].astype(str)
        )
    if "Field" in result.columns:
        result["Field"] = result["Field"].map(friendly_column_name)
    for column in columns:
        if column not in result.columns:
            result[column] = "—"
    return result.loc[:, columns]


def _audit_export_frame(
    df: pd.DataFrame,
    source_schemas: Mapping[str, Iterable[str]] | None,
    cleaning_audit: Iterable[CleaningAuditEntry | Mapping[str, Any]] | None,
) -> pd.DataFrame:
    records = audit_records(cleaning_audit or [])
    structural_records = [
        entry.as_record()
        for entry in structural_blank_audit(df, source_schemas)
    ]
    existing_keys = {
        (
            record.get("action_type"),
            str(record.get("row_index")),
            str(record.get("column")),
            str(record.get("source_file")),
        )
        for record in records
    }
    for record in structural_records:
        key = (
            record.get("action_type"),
            str(record.get("row_index")),
            str(record.get("column")),
            str(record.get("source_file")),
        )
        if key not in existing_keys:
            records.append(record)
            existing_keys.add(key)

    columns = [
        "Audit Event ID",
        "Original Source Row",
        "Record ID",
        "Source File",
        "Field",
        "Original State",
        "Action",
        "Result",
        "Method",
        "Reason",
        "Recorded At",
    ]
    if not records:
        return pd.DataFrame(columns=columns)
    raw = pd.DataFrame(records)
    def preferred(primary: str, fallback: str) -> pd.Series:
        first = raw.get(primary, pd.Series(index=raw.index, dtype=object))
        second = raw.get(fallback, pd.Series(index=raw.index, dtype=object))
        return first.where(first.notna() & first.astype(str).ne(""), second)

    frame = pd.DataFrame({
        "Audit Event ID": raw.get("audit_event_id"),
        "Original Source Row": raw.get("original_source_row"),
        "Record ID": preferred("business_record_identifier", "row_identifier"),
        "Source File": raw.get("source_file"),
        "Field": raw.get("column"),
        "Original State": raw.get("original_state"),
        "Action": raw.get("action"),
        "Result": preferred("resulting_state", "resulting_value"),
        "Method": preferred("formula_or_strategy", "strategy"),
        "Reason": raw.get("reason"),
        "Recorded At": preferred("recorded_at", "timestamp"),
    })
    parsed = pd.to_datetime(frame["Recorded At"], errors="coerce", utc=True)
    frame["Recorded At"] = [
            value.strftime("%Y-%m-%d %H:%M:%S UTC")
            if not pd.isna(value)
            else original
            for value, original in zip(parsed, frame["Recorded At"])
    ]
    return frame.loc[:, columns]


def _add_header_mappings(worksheet: Any, df: pd.DataFrame) -> None:
    """Preserve each original field name behind its friendly display header."""
    for column_index, column in enumerate(df.columns, start=1):
        worksheet.cell(row=1, column=column_index).comment = Comment(
            f"Internal field name: {column}",
            "Excel Automation Toolkit",
        )


def _apply_integrity_flags(
    worksheet: Any,
    df: pd.DataFrame,
    cleaning_report: Mapping[str, Any] | None,
) -> None:
    """Flag cleaned-data cells involved in severe relationship issues."""
    issues = list(dict(cleaning_report or {}).get("integrity_issues", []))
    if not issues:
        return
    identifier_columns = [
        column
        for column in df.columns
        if str(column).strip().lower() == "id"
        or str(column).strip().lower().endswith("_id")
    ]
    warning_fill = PatternFill("solid", fgColor="FDE68A")
    worksheet.sheet_properties.tabColor = AMBER

    for issue in issues:
        candidates = pd.Series(True, index=df.index)
        if identifier_columns:
            identifier = identifier_columns[0]
            candidates &= df[identifier].map(str).eq(
                str(issue.get("affected_record_identifier", ""))
            )
        if "source_file" in df.columns:
            candidates &= df["source_file"].map(str).eq(
                str(issue.get("source_file", ""))
            )
        involved = {
            value.strip()
            for value in str(issue.get("involved_columns", "")).split(",")
        }
        for index in df.index[candidates]:
            row_position = int(df.index.get_loc(index)) + 2
            for column_position, column in enumerate(df.columns, start=1):
                if str(column) not in involved:
                    continue
                cell = worksheet.cell(
                    row=row_position,
                    column=column_position,
                )
                cell.fill = warning_fill
                cell.comment = Comment(
                    (
                        "Severe integrity issue: "
                        f"{issue.get('expected_relationship', 'relationship check')}. "
                        f"Difference: {issue.get('difference', '—')}."
                    ),
                    "Excel Automation Toolkit",
                )


def export_to_excel(
    df: pd.DataFrame,
    filename_base: str = "",
    *,
    cleaning_report: Mapping[str, Any] | None = None,
    cleaning_audit: Iterable[CleaningAuditEntry | Mapping[str, Any]] | None = None,
    outlier_df: pd.DataFrame | None = None,
    source_schemas: Mapping[str, Iterable[str]] | None = None,
    source_files: list[str] | None = None,
) -> tuple[bytes, str]:
    """Export a professional multi-sheet Excel workbook."""
    if df.empty:
        raise ExportError(
            "Cannot export an empty DataFrame. Exporting an empty file "
            "(headers only) is blocked because it would not be useful."
        )
    if not filename_base:
        filename_base = f"cleaned_dataset_{date.today().isoformat()}"
    filename = _safe_filename(filename_base, ".xlsx")

    try:
        prepared = _prepare_export_dataframe(df, for_csv=False)
        workbook = Workbook()
        cleaned_sheet = workbook.active
        cleaned_sheet.title = EXCEL_SHEET_NAME
        _write_dataframe(cleaned_sheet, prepared, friendly_headers=True)
        _add_header_mappings(cleaned_sheet, df)
        _style_table_sheet(
            cleaned_sheet,
            autofilter_table_name="CleanedDataTable",
        )
        _apply_cleaned_data_formats(cleaned_sheet, df)
        _apply_integrity_flags(cleaned_sheet, df, cleaning_report)

        summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
        _write_cleaning_summary(
            summary_sheet,
            df,
            cleaning_report,
            source_files or [],
            source_schemas,
            cleaning_audit,
        )
        _style_table_sheet(summary_sheet)
        section_headers = {
            "Actions applied",
            "Unavailable source fields",
            "Estimated values",
            "Deterministic recoveries",
            "Unresolved problems",
            "Integrity Checks",
            "Blank status by field",
        }
        for row_index in range(2, summary_sheet.max_row + 1):
            if summary_sheet.cell(row=row_index, column=1).value in section_headers:
                if summary_sheet.cell(row=row_index, column=2).value in {
                    "Details",
                    "Result",
                    "Approved blank",
                }:
                    for cell in summary_sheet[row_index]:
                        _header_style(cell)
        summary_sheet.column_dimensions["A"].width = 31
        summary_sheet.column_dimensions["B"].width = 38
        summary_sheet.column_dimensions["C"].width = 60
        summary_sheet.column_dimensions["D"].width = 22
        for row in summary_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        _fit_wrapped_row_heights(summary_sheet)

        review_sheet = workbook.create_sheet(REVIEW_SHEET_NAME)
        review_frame = _review_export_frame(outlier_df)
        review_sheet.append([
            (
                "These values fall outside the statistical review range. "
                "They are not automatically incorrect."
            )
        ])
        review_sheet.append([])
        _write_dataframe(review_sheet, review_frame, friendly_headers=True)
        if review_frame.empty:
            empty_review = {
                column: "—"
                for column in review_frame.columns
            }
            empty_review["Product"] = "No unusual values were flagged."
            review_sheet.append([
                empty_review[column]
                for column in review_frame.columns
            ])
        _style_table_sheet(
            review_sheet,
            freeze="A4",
            autofilter_table_name="ValuesToReviewTable",
            header_row=3,
        )
        review_sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(1, review_sheet.max_column),
        )
        review_sheet["A1"].fill = PatternFill("solid", fgColor=BLUE_GRAY)
        review_sheet["A1"].font = Font(
            name="Calibri",
            size=10,
            color=MUTED,
            italic=True,
        )
        review_sheet["A1"].alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )
        review_sheet.row_dimensions[1].height = 34
        for row in review_sheet.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        review_headers = {
            str(cell.value): cell.column
            for cell in review_sheet[3]
        }
        field_column = review_headers.get("Field")
        for row_index in range(4, review_sheet.max_row + 1):
            field = (
                str(review_sheet.cell(row_index, field_column).value)
                if field_column
                else ""
            )
            for header in ("Value", "Upper Review Threshold"):
                column_index = review_headers.get(header)
                if column_index is None:
                    continue
                review_sheet.cell(
                    row=row_index,
                    column=column_index,
                ).number_format = (
                    "$#,##0.00"
                    if field in {"Unit Price", "Total"}
                    else "#,##0.00"
                )
        _fit_wrapped_row_heights(review_sheet)

        audit_sheet = workbook.create_sheet(AUDIT_SHEET_NAME)
        audit_frame = _audit_export_frame(
            df,
            source_schemas,
            cleaning_audit,
        )
        _write_dataframe(audit_sheet, audit_frame, friendly_headers=True)
        if audit_frame.empty:
            empty_audit = {
                column: "—"
                for column in audit_frame.columns
            }
            empty_audit.update({
                "action": "No cleaning actions were recorded.",
                "rows_removed": 0,
                "reason": "The exported data matches the approved dataset.",
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
            audit_sheet.append([
                empty_audit[column]
                for column in audit_frame.columns
            ])
        _style_table_sheet(
            audit_sheet,
            autofilter_table_name="CleaningAuditTable",
        )
        for row in audit_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        _fit_wrapped_row_heights(audit_sheet)

        for worksheet in workbook.worksheets:
            worksheet.auto_filter.ref = None
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0

        buffer = BytesIO()
        workbook.save(buffer)
        result = buffer.getvalue()
        integrity_passed = bool(
            dict(cleaning_report or {}).get("integrity_passed", True)
        )
        logger.info(
            "Exported %s Excel workbook: %s (%.1f KB, %d rows)",
            "business-ready" if integrity_passed else "review-required",
            filename,
            len(result) / 1024,
            len(df),
        )
        return result, filename
    except ExportError:
        raise
    except Exception as exc:
        logger.exception("Excel export failed")
        raise ExportError(f"Excel export failed: {exc}") from exc


def export_to_csv(
    df: pd.DataFrame,
    filename_base: str = "",
) -> tuple[bytes, str]:
    """Export a simple UTF-8-with-BOM CSV matching the cleaned data."""
    if df.empty:
        raise ExportError(
            "Cannot export an empty DataFrame. Exporting an empty file "
            "(headers only) is blocked because it would not be useful."
        )
    if not filename_base:
        filename_base = f"cleaned_dataset_{date.today().isoformat()}"
    filename = _safe_filename(filename_base, ".csv")

    try:
        prepared = _prepare_export_dataframe(df, for_csv=True)
        text = prepared.to_csv(index=False, lineterminator="\n", na_rep="")
        result = text.encode("utf-8-sig")
        logger.info(
            "Exported CSV: %s (%.1f KB, %d rows, %d cols)",
            filename,
            len(result) / 1024,
            len(df),
            len(df.columns),
        )
        return result, filename
    except ExportError:
        raise
    except Exception as exc:
        logger.exception("CSV export failed")
        raise ExportError(f"CSV export failed: {exc}") from exc


def export_audit_to_csv(
    cleaning_audit: Iterable[CleaningAuditEntry | Mapping[str, Any]],
    filename_base: str = "cleaning_audit",
) -> tuple[bytes, str]:
    """Export the structured cleaning audit as an optional companion CSV."""
    records = audit_records(cleaning_audit)
    filename = _safe_filename(filename_base, ".csv")
    frame = pd.DataFrame(records)
    text = frame.to_csv(index=False, lineterminator="\n", na_rep="")
    return text.encode("utf-8-sig"), filename


def export_missingness_summary(
    df: pd.DataFrame,
    source_schemas: Mapping[str, Iterable[str]] | None,
) -> pd.DataFrame:
    """Expose the exporter's reconciled missingness view for verification."""
    return classify_missing_values(df, source_schemas)
